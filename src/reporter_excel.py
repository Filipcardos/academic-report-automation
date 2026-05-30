"""
Reporter Excel — gera relatório .xlsx com múltiplas abas e gráficos.
"""

import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

log = logging.getLogger(__name__)

# ── Paleta ──────────────────────────────────────────────────────────────
BLUE_DARK = "1F3864"
BLUE_MID = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GREEN = "70AD47"
RED = "C00000"
GRAY = "F2F2F2"
WHITE = "FFFFFF"


def _header_style(cell, bg=BLUE_DARK, fg=WHITE, size=11, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_table(ws, headers, rows, start_row=1, start_col=1):
    """Escreve cabeçalho + linhas com zebra striping."""
    for ci, h in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=ci, value=h)
        _header_style(cell, bg=BLUE_MID)

    for ri, row in enumerate(rows, start_row + 1):
        bg = GRAY if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row, start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()


# ── Aba: Dashboard ──────────────────────────────────────────────────────
def _sheet_dashboard(wb, data):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Título
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "RELATÓRIO ACADÊMICO — INDICADORES GERAIS"
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtítulo data
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Arial", size=10, color="666666")
    c.alignment = Alignment(horizontal="center")

    # KPIs
    kpis = [
        ("Total de Alunos", data["total_alunos"]),
        ("Aprovados", data["total_aprovados"]),
        ("Reprovados", data["total_reprovados"]),
        ("Taxa de Aprovação", f"{data['taxa_aprovacao']}%"),
        ("Média Geral", data["media_geral"]),
    ]
    for i, (label, val) in enumerate(kpis, 4):
        cl = ws.cell(row=i, column=1, value=label)
        cv = ws.cell(row=i, column=2, value=val)
        cl.font = Font(name="Arial", bold=True, size=11)
        cl.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        cv.font = Font(name="Arial", bold=True, size=12, color=BLUE_DARK)
        cv.fill = PatternFill("solid", fgColor=WHITE)
        cv.alignment = Alignment(horizontal="center")
        ws.row_dimensions[i].height = 22


# ── Aba: Por Curso ──────────────────────────────────────────────────────
def _sheet_por_curso(wb, data):
    ws = wb.create_sheet("Por Curso")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [22, 10, 12, 12, 18, 10]):
        ws.column_dimensions[col].width = w

    headers = [
        "Curso",
        "Total",
        "Aprovados",
        "Reprovados",
        "Taxa Aprovação (%)",
        "Média",
    ]
    rows = [
        (
            r["curso"],
            r["total"],
            r["aprovados"],
            r["reprovados"],
            r["taxa_aprovacao"],
            r["media"],
        )
        for r in data["por_curso"]
    ]
    _write_table(ws, headers, rows, start_row=2)

    # Título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Desempenho por Curso"
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.title = "Taxa de Aprovação por Curso"
    chart.y_axis.title = "%"
    chart.style = 10

    data_ref = Reference(ws, min_col=5, min_row=2, max_row=2 + len(rows))
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "H2")


# ── Aba: Por Disciplina ─────────────────────────────────────────────────
def _sheet_por_disciplina(wb, data):
    ws = wb.create_sheet("Por Disciplina")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [22, 10, 12, 12, 18, 10]):
        ws.column_dimensions[col].width = w

    headers = [
        "Disciplina",
        "Total",
        "Aprovados",
        "Reprovados",
        "Taxa Aprovação (%)",
        "Média",
    ]
    rows = [
        (
            r["disciplina"],
            r["total"],
            r["aprovados"],
            r["reprovados"],
            r["taxa_aprovacao"],
            r["media"],
        )
        for r in data["por_disciplina"]
    ]
    _write_table(ws, headers, rows, start_row=2)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Desempenho por Disciplina"
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


# ── Aba: Alunos ─────────────────────────────────────────────────────────
def _sheet_alunos(wb, data):
    ws = wb.create_sheet("Alunos")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [8, 22, 28, 18, 8, 10, 12]):
        ws.column_dimensions[col].width = w

    headers = ["ID", "Nome", "E-mail", "Curso", "Período", "Média", "Situação"]
    rows = [
        (
            a["id_aluno"],
            a["nome"],
            a["email"],
            a["curso"],
            a["periodo"],
            a["media"],
            a["situacao"],
        )
        for a in sorted(data["alunos"], key=lambda x: x["nome"])
    ]
    _write_table(ws, headers, rows, start_row=2)

    # Colorir situação
    for ri, row in enumerate(rows, 3):
        cell = ws.cell(row=ri, column=7)
        if cell.value == "Aprovado":
            cell.font = Font(name="Arial", bold=True, color=GREEN, size=10)
        else:
            cell.font = Font(name="Arial", bold=True, color=RED, size=10)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Lista de Alunos"
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.auto_filter.ref = f"A2:G{len(rows) + 2}"


# ── Entry point ─────────────────────────────────────────────────────────
def generate_excel(data: dict, path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão

    _sheet_dashboard(wb, data)
    _sheet_por_curso(wb, data)
    _sheet_por_disciplina(wb, data)
    _sheet_alunos(wb, data)

    wb.save(path)
    log.info(f"  Excel salvo: {path}")
